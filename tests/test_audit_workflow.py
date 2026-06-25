from __future__ import annotations

import csv
import json

from openpyxl import load_workbook
import pytest

from claw_easa.audit.export import export_report_csv, export_report_json, export_report_xlsx
from claw_easa.audit.schema import load_report, validate_report
from claw_easa.audit.storage import fetch_finding, fetch_report, import_report, list_reports
from claw_easa.config import Settings, reset_settings
from claw_easa.db import Database
from claw_easa.db.migrations import MigrationRunner


@pytest.fixture
def audit_settings(tmp_path):
    reset_settings()
    settings = Settings(data_dir=str(tmp_path), db_file="audit.db")
    db = Database(settings=settings)
    db.open()
    MigrationRunner(db).init_schema()
    db.close()
    yield settings
    reset_settings()


@pytest.fixture
def sample_report_v1():
    return {
        "schema_version": "1.0",
        "report_id": "AUD-20260422-0001",
        "report_name": "OM-A pilot audit",
        "manual_name": "OM-A",
        "manual_version_date": "Rev. 12 / 2026-03-01",
        "entity_scope": "ASLB",
        "created_at": "2026-04-22T12:00:00Z",
        "findings": [
            {
                "finding_id": "AUD-20260422-0101",
                "manual_name": "OM-A",
                "manual_section_paragraph": "4.2.1",
                "manual_version_date": "Rev. 12 / 2026-03-01",
                "entity_scope": "ASLB",
                "applicable_easa_references": ["ORO.FTL.110(a)"],
                "source_hierarchy_notes": ["IR is binding", "No AMC/GM required"],
                "manual_excerpt": "The company publishes duty rosters in advance.",
                "easa_excerpts": [
                    "An operator shall publish duty rosters sufficiently in advance to provide the opportunity for crew members to plan adequate rest."
                ],
                "assessment": "The manual covers the core obligation but does not define the advance notice.",
                "compliance_score": 4,
                "severity": "Low",
                "confidence": "High",
                "gap_types": ["wording / editorial weakness"],
                "recommendation": "Specify the minimum publication lead time.",
                "review_status": "Proposed",
            }
        ],
    }


@pytest.fixture
def sample_report_v2():
    return {
        "schema_version": "1.0",
        "report_id": "AUD-20260423-0001",
        "report_name": "OM-A pilot audit — updated",
        "manual_name": "OM-A",
        "manual_version_date": "Rev. 13 / 2026-04-15",
        "entity_scope": "ASLB",
        "created_at": "2026-04-23T12:00:00Z",
        "findings": [
            {
                "finding_id": "AUD-20260422-0101",
                "manual_name": "OM-A",
                "manual_section_paragraph": "4.2.1",
                "manual_version_date": "Rev. 13 / 2026-04-15",
                "entity_scope": "ASLB",
                "applicable_easa_references": ["ORO.FTL.110(a)"],
                "source_hierarchy_notes": ["IR is binding", "No AMC/GM required"],
                "manual_excerpt": "The company publishes duty rosters at least 14 days in advance.",
                "easa_excerpts": [
                    "An operator shall publish duty rosters sufficiently in advance to provide the opportunity for crew members to plan adequate rest."
                ],
                "assessment": "The manual now defines an explicit publication lead time.",
                "compliance_score": 5,
                "severity": "Low",
                "confidence": "High",
                "gap_types": [],
                "recommendation": "Keep the lead time controlled in the next manual revision.",
                "review_status": "Approved",
            }
        ],
    }


class TestAuditSchema:
    def test_validate_report_canonicalizes(self, sample_report_v1):
        report = validate_report(sample_report_v1)
        assert report["report_id"] == "AUD-20260422-0001"
        assert report["findings"][0]["compliance_score"] == 4

    def test_load_report_roundtrip(self, sample_report_v1, tmp_path):
        path = tmp_path / "report.json"
        path.write_text(json.dumps(sample_report_v1), encoding="utf-8")
        report = load_report(path)
        assert report["manual_name"] == "OM-A"
        assert report["findings"][0]["finding_id"] == "AUD-20260422-0101"


class TestAuditPersistence:
    def test_import_and_fetch_report_and_finding(self, audit_settings, sample_report_v1, sample_report_v2):
        db = Database(settings=audit_settings)
        db.open()
        try:
            imported = import_report(db, sample_report_v1, source_path="/tmp/report-v1.json")
            assert imported["report_id"] == sample_report_v1["report_id"]

            imported = import_report(db, sample_report_v2, source_path="/tmp/report-v2.json")
            assert imported["report_id"] == sample_report_v2["report_id"]

            fetched = fetch_report(db, sample_report_v1["report_id"])
            assert fetched["report_name"] == "OM-A pilot audit"

            finding = fetch_finding(db, "AUD-20260422-0101")
            assert finding["finding_id"] == "AUD-20260422-0101"
            assert finding["latest_revision_number"] == 2
            assert len(finding["revisions"]) == 2
            assert finding["latest_revision"]["compliance_score"] == 5
            assert finding["latest_revision"]["review_status"] == "Approved"
            assert finding["latest_revision"]["evidence"][0]["evidence_kind"] == "manual"

            reports = list_reports(db)
            assert len(reports) == 2
            assert reports[0]["report_id"] == sample_report_v2["report_id"]
            assert reports[0]["finding_count"] == 1
        finally:
            db.close()


class TestAuditExports:
    def test_export_json(self, sample_report_v1, tmp_path):
        out = export_report_json(sample_report_v1, tmp_path / "report.export.json")
        loaded = json.loads(out.read_text(encoding="utf-8"))
        assert loaded["report_id"] == sample_report_v1["report_id"]
        assert loaded["findings"][0]["finding_id"] == "AUD-20260422-0101"

    def test_export_csv(self, sample_report_v1, tmp_path):
        out = export_report_csv(sample_report_v1, tmp_path / "report.csv")
        with out.open(encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 1
        assert rows[0]["finding_id"] == "AUD-20260422-0101"
        assert rows[0]["revision_number"] == "1"
        assert "ORO.FTL.110(a)" in rows[0]["applicable_easa_references"]

    def test_export_xlsx(self, sample_report_v1, tmp_path):
        out = export_report_xlsx(sample_report_v1, tmp_path / "report.xlsx")
        wb = load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "Findings" in wb.sheetnames
        ws = wb["Findings"]
        assert ws["A2"].value == "AUD-20260422-0101"
        assert ws["B2"].value == 1
        assert ws["C2"].value == 1
